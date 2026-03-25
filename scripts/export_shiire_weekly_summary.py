#!/usr/bin/env python3
"""
仕入れ先(工程)別 仕入金額 週集計エクセル出力

要件:
- 仕入伝票(denpyou_mr_id=2) の内訳10を集計
- 対象工程: WM / WS / G(合撚) / IT(イタリー) / KR(仮撚) / JNB(整経) / OR(製織)
- 過去3ヶ月(デフォルト)を月別ブロックで横並び
- 右側が最新月
- 月内週は月曜起点ベース、月末は月末日で打ち切り
  - 月初が日曜日の場合は 1日だけの週を避けるため、1日〜翌週日曜を第1区間とする

実行例:
  python3 scripts/export_shiire_weekly_summary.py
  python3 scripts/export_shiire_weekly_summary.py --as-of 2026-03-02 --months 3
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import os
from decimal import Decimal
from typing import Dict, List, Sequence, Tuple

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROCESS_ORDER: List[Tuple[str, str]] = [
    ("WM", "WM"),
    ("WS", "WS"),
    ("G", "合撚 (G)"),
    ("IT", "イタリー (IT)"),
    ("KR", "仮撚 (KR)"),
    ("JNB", "整経 (JNB)"),
    ("OR", "製織 (OR)"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="仕入れ週集計エクセル出力")
    parser.add_argument(
        "--as-of",
        default=(dt.date.today() - dt.timedelta(days=1)).isoformat(),
        help="集計対象の最終日 YYYY-MM-DD (デフォルト: 前日)",
    )
    parser.add_argument("--months", type=int, default=3, help="対象月数(右端が最新)")
    parser.add_argument("--host", default=os.getenv("SFN4_DB_HOST", "192.168.1.5"))
    parser.add_argument("--port", type=int, default=int(os.getenv("SFN4_DB_PORT", "3306")))
    parser.add_argument("--db", default=os.getenv("SFN4_DB_NAME", "sfn4"))
    parser.add_argument("--user", default=os.getenv("SFN4_DB_USER", "erphalcon"))
    parser.add_argument("--password", default=os.getenv("SFN4_DB_PASSWORD", "erphalcon"))
    parser.add_argument(
        "--amount-column",
        choices=["zeinukigaku", "kingaku"],
        default="zeinukigaku",
        help="集計金額列(デフォルト:税抜金額)",
    )
    parser.add_argument("--output", default="", help="出力先xlsxパス")
    return parser.parse_args()


def month_start(d: dt.date) -> dt.date:
    return d.replace(day=1)


def month_end(d: dt.date) -> dt.date:
    last = calendar.monthrange(d.year, d.month)[1]
    return d.replace(day=last)


def add_months(d: dt.date, months: int) -> dt.date:
    year = d.year + (d.month - 1 + months) // 12
    month = (d.month - 1 + months) % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return dt.date(year, month, day)


def build_target_months(as_of: dt.date, months: int) -> List[dt.date]:
    base = month_start(as_of)
    return [add_months(base, -i) for i in reversed(range(months))]


def build_month_periods(target_month: dt.date, cap_date: dt.date) -> List[Tuple[dt.date, dt.date]]:
    """
    月内の区間を作る。
    - 基本は月曜起点の週区切り
    - 月初が日曜なら 1日だけの区間を避けるため翌週日曜までを第1区間にする
    - 最終区間は月末(cap_date)で打ち切り
    """
    start = month_start(target_month)
    end = min(month_end(target_month), cap_date)
    if start > end:
        return []

    periods: List[Tuple[dt.date, dt.date]] = []

    first_sunday = start + dt.timedelta(days=(6 - start.weekday()) % 7)
    if first_sunday == start:
        first_end = min(start + dt.timedelta(days=7), end)
    else:
        first_end = min(first_sunday, end)

    periods.append((start, first_end))
    cursor = first_end + dt.timedelta(days=1)

    while cursor <= end:
        period_end = min(cursor + dt.timedelta(days=6), end)
        periods.append((cursor, period_end))
        cursor = period_end + dt.timedelta(days=1)

    return periods


def fetch_daily_amounts(
    conn: pymysql.connections.Connection,
    from_date: dt.date,
    to_date: dt.date,
    amount_column: str,
) -> Dict[Tuple[str, dt.date], Decimal]:
    process_codes = [p[0] for p in PROCESS_ORDER]
    placeholders = ",".join(["%s"] * len(process_codes))

    sql = f"""
        SELECT
            a.souko_mr_cd AS souko_mr_cd,
            DATE(b.hakkoubi) AS hakkoubi,
            COALESCE(SUM(a.{amount_column}), 0) AS amount
        FROM gyoumu_meisai_dts AS a
        JOIN gyoumu_dts AS b ON b.id = a.gyoumu_dt_id AND b.denpyou_mr_id = 2
        WHERE a.utiwake_kbn_cd = '10'
          AND a.souko_mr_cd IN ({placeholders})
          AND b.hakkoubi BETWEEN %s AND %s
        GROUP BY a.souko_mr_cd, DATE(b.hakkoubi)
    """

    params: List[object] = process_codes + [from_date.isoformat(), to_date.isoformat()]
    result: Dict[Tuple[str, dt.date], Decimal] = {}
    with conn.cursor() as cur:
        cur.execute(sql, params)
        for code, ymd, amount in cur.fetchall():
            if ymd is None:
                continue
            result[(str(code), ymd)] = Decimal(str(amount))
    return result


def sum_period(
    daily_map: Dict[Tuple[str, dt.date], Decimal],
    code: str,
    start: dt.date,
    end: dt.date,
) -> Decimal:
    total = Decimal("0")
    d = start
    while d <= end:
        total += daily_map.get((code, d), Decimal("0"))
        d += dt.timedelta(days=1)
    return total


def period_label(start: dt.date, end: dt.date) -> str:
    return f"{start.strftime('%y/%-m/%-d')}~{end.strftime('%-m/%-d')}"


def output_path(as_of: dt.date, explicit: str) -> str:
    if explicit:
        return explicit
    out_dir = "/Volumes/www/html/groupware/exports"
    os.makedirs(out_dir, exist_ok=True)
    return os.path.join(out_dir, f"shiire_weekly_summary_{as_of.strftime('%Y%m%d')}.xlsx")


def write_excel(
    out_file: str,
    as_of: dt.date,
    target_months: Sequence[dt.date],
    month_periods: Dict[dt.date, List[Tuple[dt.date, dt.date]]],
    daily_map: Dict[Tuple[str, dt.date], Decimal],
    amount_column: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "週集計"

    total_cols = 1 + len(target_months) * 2

    ws.cell(row=1, column=1, value="仕入れ先別 仕入金額 週集計 (内訳10)")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=2, column=1, value=f"基準日: {as_of.isoformat()} / 金額列: {amount_column}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    ws.cell(row=3, column=1, value="工程")
    for i, m in enumerate(target_months):
        start_col = 2 + i * 2
        ws.cell(row=3, column=start_col, value=m.strftime("%Y/%m"))
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + 1)
        ws.cell(row=4, column=start_col, value="週")
        ws.cell(row=4, column=start_col + 1, value="仕入金額")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    title_fill = PatternFill("solid", fgColor="BDD7EE")
    total_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="777777")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in (1, 2):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = title_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(bold=True)

    for r in (3, 4):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    row = 5
    max_rows = max(len(p) for p in month_periods.values()) if month_periods else 0

    for code, label in PROCESS_ORDER:
        block_rows = max_rows if max_rows > 0 else 1
        ws.cell(row=row, column=1, value=label)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + block_rows - 1, end_column=1)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).font = Font(bold=True)

        month_totals: Dict[dt.date, Decimal] = {m: Decimal("0") for m in target_months}

        for i in range(block_rows):
            for m_idx, m in enumerate(target_months):
                start_col = 2 + m_idx * 2
                periods = month_periods[m]
                week_cell = ws.cell(row=row + i, column=start_col)
                amount_cell = ws.cell(row=row + i, column=start_col + 1)

                if i < len(periods):
                    p_start, p_end = periods[i]
                    amt = sum_period(daily_map, code, p_start, p_end)
                    week_cell.value = period_label(p_start, p_end)
                    amount_cell.value = int(amt)
                    amount_cell.number_format = "#,##0"
                    month_totals[m] += amt

                week_cell.alignment = Alignment(horizontal="center", vertical="center")
                amount_cell.alignment = Alignment(horizontal="right", vertical="center")

        total_row = row + block_rows
        ws.cell(row=total_row, column=1, value="計")
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = total_fill

        for m_idx, m in enumerate(target_months):
            start_col = 2 + m_idx * 2
            ws.cell(row=total_row, column=start_col, value="計")
            ws.cell(row=total_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=total_row, column=start_col).font = Font(bold=True)
            ws.cell(row=total_row, column=start_col).fill = total_fill

            total_cell = ws.cell(row=total_row, column=start_col + 1, value=int(month_totals[m]))
            total_cell.number_format = "#,##0"
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_cell.font = Font(bold=True)
            total_cell.fill = total_fill

        for rr in range(row, total_row + 1):
            for cc in range(1, total_cols + 1):
                ws.cell(row=rr, column=cc).border = border

        row = total_row + 1

    ws.column_dimensions["A"].width = 14
    for i in range(len(target_months)):
        col_week = chr(ord("B") + i * 2)
        col_amt = chr(ord("C") + i * 2)
        ws.column_dimensions[col_week].width = 16
        ws.column_dimensions[col_amt].width = 14

    ws.freeze_panes = "B5"
    wb.save(out_file)


def main() -> None:
    args = parse_args()
    as_of = dt.date.fromisoformat(args.as_of)

    target_months = build_target_months(as_of, args.months)
    month_periods: Dict[dt.date, List[Tuple[dt.date, dt.date]]] = {}
    for m in target_months:
        cap = min(month_end(m), as_of if m == month_start(as_of) else month_end(m))
        month_periods[m] = build_month_periods(m, cap)

    data_from = month_start(target_months[0])
    data_to = min(month_end(target_months[-1]), as_of)

    conn = pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.db,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.Cursor,
    )
    try:
        daily_map = fetch_daily_amounts(conn, data_from, data_to, args.amount_column)
    finally:
        conn.close()

    out_file = output_path(as_of, args.output)
    write_excel(out_file, as_of, target_months, month_periods, daily_map, args.amount_column)
    print(out_file)


if __name__ == "__main__":
    main()
